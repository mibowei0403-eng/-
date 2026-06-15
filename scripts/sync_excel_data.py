import json
import math
import shutil
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "business-dashboard-data.json"
RENTAL_FILE = Path("/Users/mibowei/Library/Mobile Documents/com~apple~CloudDocs/AwStudio/租赁统计表.xlsx")
FINANCE_FILE = Path("/Users/mibowei/Library/Mobile Documents/com~apple~CloudDocs/AwStudio/财务收入支出表.xlsx")


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def text(value):
    value = clean(value)
    if value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    value = clean(value)
    if value == "":
        return 0
    try:
        if isinstance(value, str):
            value = value.replace(",", "")
        n = float(value)
        if math.isnan(n):
            return 0
        return int(n) if n.is_integer() else round(n, 2)
    except Exception:
        return 0


def date_text(value):
    value = clean(value)
    if value == "":
        return ""
    if isinstance(value, datetime):
        return f"{value.year}/{value.month}/{value.day}"
    if isinstance(value, (int, float)) and value > 20000:
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=float(value))
        return f"{dt.year}/{dt.month}/{dt.day}"
    return str(value).strip()


def normalize_spec(value):
    return " ".join(text(value).split())


def meaningful(values):
    return any(clean(v) != "" for v in values)


def rental_order_from_row(row, excel_row, order_id):
    return {
        "id": order_id,
        "row": excel_row,
        "no": text(row[0]),
        "deviceCode": text(row[1]),
        "model": normalize_spec(row[2]),
        "cost": number(row[3]),
        "cycleCost": number(row[4]),
        "startDate": date_text(row[5]),
        "returnDate": date_text(row[6]),
        "expectedMonths": number(row[7]) if clean(row[7]) != "" else "",
        "monthlyRent": number(row[8]),
        "currentMonths": number(row[9]) if clean(row[9]) != "" else "",
        "collected": number(row[10]),
        "status": text(row[11]),
        "customer": text(row[12]),
        "phone": text(row[13]),
        "idCard": text(row[14]),
        "accessories": text(row[15]),
        "note": text(row[16]),
        "rentTime": text(row[17]),
    }


def parse_rental_orders():
    wb = load_workbook(RENTAL_FILE, data_only=True, read_only=True)
    orders = []

    ws = wb["统计总表"]
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row <= 2:
            continue
        if not meaningful(row[:18]):
            continue
        if clean(row[0]) == "" and clean(row[1]) == "" and clean(row[2]) == "":
            continue
        if text(row[11]) == "空置" and text(row[12]) == "" and date_text(row[5]) == "":
            continue
        orders.append(rental_order_from_row(row, excel_row, f"rent-{excel_row}"))

    if "张欣" in wb.sheetnames:
        ws = wb["张欣"]
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not meaningful(row[:18]):
                continue
            if clean(row[0]) == "" and clean(row[1]) == "" and clean(row[2]) == "":
                continue
            order = rental_order_from_row(row, idx, f"rent-zhangxin-{text(row[0]) or idx}")
            order["status"] = "坏单"
            order["badDebtId"] = "bad-debt-zhangxin"
            orders.append(order)

    return orders


def parse_idle_devices_from_rental():
    wb = load_workbook(RENTAL_FILE, data_only=True, read_only=True)
    ws = wb["统计总表"]
    devices = []
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row <= 2:
            continue
        if text(row[11]) != "空置" or text(row[12]) or date_text(row[5]):
            continue
        code = text(row[1])
        if not code:
            continue
        devices.append({
            "id": code,
            "code": code,
            "brandModel": "台式机",
            "spec": normalize_spec(row[2]),
            "cost": number(row[3]),
            "bookCost": number(row[3]),
            "status": "空置",
            "depositFree": "",
            "rent": number(row[8]),
            "rentedMonths": "",
            "collected": number(row[10]),
            "paybackProgress": "",
            "paidBack": "否",
            "currentCustomer": "",
        })
    return devices


def parse_income():
    wb = load_workbook(FINANCE_FILE, data_only=True, read_only=True)
    ws = wb["收入明细"]
    rows = []
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row <= 8:
            continue
        if clean(row[3]) == "" and clean(row[4]) == "" and clean(row[7]) == "":
            continue
        if not isinstance(clean(row[3]), (int, float)):
            continue
        rows.append({
            "id": f"inc-{excel_row}",
            "row": excel_row,
            "date": date_text(row[4]),
            "category": text(row[5]),
            "account": text(row[6]),
            "amount": number(row[7]),
            "customer": text(row[8]),
            "summary": text(row[9]),
            "entryDate": date_text(row[11]),
        })
    return rows


def parse_expense():
    wb = load_workbook(FINANCE_FILE, data_only=True, read_only=True)
    ws = wb["支出明细"]
    rows = []
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row <= 8:
            continue
        if clean(row[3]) == "" and clean(row[4]) == "" and clean(row[7]) == "":
            continue
        if not isinstance(clean(row[3]), (int, float)):
            continue
        rows.append({
            "id": f"exp-{excel_row}",
            "row": excel_row,
            "date": date_text(row[4]),
            "category": text(row[5]),
            "account": text(row[6]),
            "amount": number(row[7]),
            "handler": text(row[8]),
            "summary": text(row[9]),
            "entryDate": date_text(row[11]),
        })
    return rows


def parse_loans():
    wb = load_workbook(FINANCE_FILE, data_only=True, read_only=True)
    ws = wb["投资款"]
    loans = []
    for excel_row, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if excel_row <= 3:
            continue
        if text(row[0]) == "合计":
            break
        if clean(row[1]) == "" and clean(row[2]) == "" and clean(row[3]) == "":
            continue
        principal = number(row[3])
        remaining = number(row[4])
        repaid = number(row[5])
        interest = number(row[6])
        if not principal and not remaining and not repaid:
            continue

        payment_values = []
        interest_values = []
        for col in range(7, min(len(row), 31), 2):
            payment = number(row[col])
            intr = number(row[col + 1]) if col + 1 < len(row) else 0
            if payment:
                payment_values.append(payment)
            if intr:
                interest_values.append(intr)

        category = text(row[1])
        name = text(row[2])
        repayment_type = "installment"
        repayment_note = ""
        paid_terms = len(payment_values) if repaid else 0
        terms = 12 if "12" in name else max(12, paid_terms)
        monthly_payment = round(principal / terms, 2) if principal and terms else 0
        monthly_interest = round(sum(interest_values) / len(interest_values), 2) if interest_values else 0
        if payment_values:
            monthly_payment = round(sum(payment_values) / len(payment_values), 2)

        if category == "个人" and name == "陈晓玲":
            repayment_type = "flexible"
            repayment_note = "对象周转款，不计入固定月供，有钱时归还本金。"
            terms = ""
            monthly_payment = 0
            monthly_interest = 0
        elif category == "个人" and name == "信用卡" and remaining == 4000:
            repayment_type = "short_term"
            repayment_note = "短期待还款，不按分期月供计算，近期一次性处理。"
            terms = ""
            monthly_payment = 0
            monthly_interest = 0
        elif category == "贷款" and name == "招商银行" and remaining > 0 and principal >= 100000:
            repayment_type = "interest_only"
            repayment_note = "先息后本，每月只计利息，本金到期或有资金时归还。"
            terms = ""
            monthly_payment = 0

        loans.append({
            "id": f"loan-{excel_row}",
            "date": date_text(row[0]),
            "category": category,
            "name": name,
            "principal": principal,
            "remainingPrincipal": remaining,
            "repaidPrincipal": repaid,
            "paidInterest": interest or round(sum(interest_values), 2),
            "paidTerms": paid_terms,
            "terms": terms,
            "monthlyPayment": monthly_payment,
            "monthlyInterest": monthly_interest,
            "repaymentType": repayment_type,
            "repaymentNote": repayment_note,
        })
    return loans


def apply_bad_debt_rules(data):
    zhang_ids = []
    zhang_codes = set()
    for order in data["rentalOrders"]:
        if text(order.get("customer")) == "张欣":
            order["status"] = "坏单"
            order["badDebtId"] = "bad-debt-zhangxin"
            zhang_ids.append(order["id"])
            if order.get("deviceCode"):
                zhang_codes.add(order["deviceCode"])

    chen_ids = []
    for order in data["rentalOrders"]:
        if text(order.get("customer")) == "陈鹏远":
            order["status"] = "坏单"
            order["badDebtId"] = "bad-debt-chenpengyuan"
            chen_ids.append(order["id"])

    for device in data.get("devices", []):
        if device.get("code") in zhang_codes:
            device["status"] = "坏单"
            device["currentCustomer"] = "张欣"

    bad_debts = data.setdefault("badDebts", [])
    by_customer = {text(item.get("customer")): item for item in bad_debts}

    zhang = by_customer.get("张欣")
    if zhang:
        zhang["rentalOrderId"] = zhang_ids[0] if zhang_ids else zhang.get("rentalOrderId", "")
        zhang["relatedRentalOrderIds"] = zhang_ids
        zhang["collectedAmount"] = max(number(zhang.get("collectedAmount")), max((number(o.get("collected")) for o in data["rentalOrders"] if text(o.get("customer")) == "张欣"), default=0))
        zhang["updatedAt"] = "2026-06-15"

    chen = by_customer.get("陈鹏远")
    if chen and chen_ids:
        chen["rentalOrderId"] = chen_ids[0]
        chen["relatedRentalOrderIds"] = chen_ids


def apply_idle_devices_from_rental(data):
    idle_rows = parse_idle_devices_from_rental()
    devices = data.setdefault("devices", [])
    by_code = {device.get("code"): device for device in devices}
    for row in idle_rows:
        device = by_code.get(row["code"])
        if not device:
            device = row.copy()
            devices.append(device)
            by_code[row["code"]] = device
            continue
        device["status"] = "空置"
        device["currentCustomer"] = ""
        if row.get("spec"):
            device["spec"] = row["spec"]
        if number(row.get("cost")):
            device["cost"] = row["cost"]
            device["bookCost"] = row["bookCost"]
        if number(row.get("rent")):
            device["rent"] = row["rent"]


def main():
    data = json.loads(DATA_FILE.read_text(encoding="utf-8"))
    backup = DATA_FILE.with_name(f"business-dashboard-data.before-excel-sync-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json")
    shutil.copy2(DATA_FILE, backup)

    data["rentalOrders"] = parse_rental_orders()
    data["income"] = parse_income()
    data["expense"] = parse_expense()
    data["loans"] = parse_loans()
    data.setdefault("managedDevices", [])
    data.setdefault("managedDeviceEvents", [])
    data.setdefault("customers", [])
    data.setdefault("badDebts", [])
    data["meta"] = {
        **data.get("meta", {}),
        "syncedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFiles": {
            **data.get("meta", {}).get("sourceFiles", {}),
            "rental": str(RENTAL_FILE),
            "finance": str(FINANCE_FILE),
        },
    }
    apply_bad_debt_rules(data)
    apply_idle_devices_from_rental(data)

    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "backup": str(backup),
        "rentalOrders": len(data["rentalOrders"]),
        "income": len(data["income"]),
        "expense": len(data["expense"]),
        "loans": len(data["loans"]),
        "zhangXinBadDebtOrders": sum(1 for o in data["rentalOrders"] if text(o.get("customer")) == "张欣" and o.get("status") == "坏单"),
        "regularCurrentOrders": sum(1 for o in data["rentalOrders"] if o.get("status") == "租赁中" and text(o.get("customer")) != "张欣"),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
